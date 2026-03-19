import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

output_path = r"w:\The Testing Acedamy\Antigravity\My Projects\Project_01_VWO\Test_ Delivarables_Outcome\VWO_Test_Metrics.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Test Metrics"

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def apply_border(cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

# TABLE 1
ws.merge_cells('A1:B1')
ws['A1'] = "Test Metrics"
ws['A1'].font = Font(bold=True, color="FFFFFF", size=12)
ws['A1'].fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid") # Light Blue
ws['A1'].alignment = Alignment(horizontal="center")
ws['A1'].border = thin_border
ws['B1'].border = thin_border

table1_data = [
    ("Number of Requirements", 10),
    ("Average Number of Test Cases Written Per Requirement", 2.7),
    ("Total Number of Test Cases Written for All Requirements", 27),
    ("Total Number of Test Cases Executed", 27),
    ("Number of Test Cases Passed", 12),
    ("Number of Test Cases Failed", 15),
    ("Number of Test Cases Blocked", 0),
    ("Number of Test Cases Unexecuted", 2)
]

for i, (cat, val) in enumerate(table1_data):
    row = i + 2
    ws.cell(row=row, column=1, value=cat).border = thin_border
    ws.cell(row=row, column=2, value=val).border = thin_border

ws['B6'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Green
ws['B7'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Red

# Categories Table (appended to Table 1)
cat_start = 10
ws.cell(row=cat_start, column=1, value="Category").font = Font(bold=True)
ws.cell(row=cat_start, column=1).border = thin_border
ws.cell(row=cat_start, column=2, value="Count").font = Font(bold=True)
ws.cell(row=cat_start, column=2).fill = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid") # Orange
ws.cell(row=cat_start, column=2).border = thin_border

categories = [
    ("Accessibility", 1),
    ("Functional", 9),
    ("Integration", 1),
    ("Negative", 1),
    ("Performance", 1),
    ("Security", 2)
]

for i, (cat, val) in enumerate(categories):
    row = cat_start + 1 + i
    ws.cell(row=row, column=1, value=cat).border = thin_border
    ws.cell(row=row, column=2, value=val).border = thin_border


# TABLE 2
ws.merge_cells('D4:E4')
ws['D4'] = "Test Metrics"
ws['D4'].font = Font(bold=True, color="FFFFFF", size=12)
ws['D4'].fill = PatternFill(start_color="9900FF", end_color="9900FF", fill_type="solid") # Purple
ws['D4'].alignment = Alignment(horizontal="center")
ws['D4'].border = thin_border
ws['E4'].border = thin_border

table2_data = [
    ("Percentage of Test Cases Executed", 100),
    ("Percentage of Test Cases Not Executed", 0),
    ("Percentage of Test Cases Passed", 44.4),
    ("Percentage of Test Cases Failed", 55.6),
    ("Percentage of Test Cases Blocked", 0)
]

for i, (cat, val) in enumerate(table2_data):
    row = 5 + i
    ws.cell(row=row, column=4, value=cat).border = thin_border
    ws.cell(row=row, column=5, value=val).border = thin_border


# FORMULAS
formulas = [
    "% of Test cases Executed:",
    "(Number of Test Cases Executed / Total Number of Test Cases Written ) * 100",
    "▪ % of test cases NOT executed:",
    "(Number of Test Cases Unexecuted / Total Number of Test Cases Written) * 100",
    "▪ % Test cases passed",
    "(Number of Test Cases Passed / Total Test Cases Executed) * 100",
    "▪ % Test cases failed",
    "(Number of Test Cases Failed / Total Test Cases Executed) * 100",
    "▪ %Test cases blocked",
    "(Number of test cases blocked / Total Test cases executed ) * 100"
]

for i, f in enumerate(formulas):
    row = 5 + i
    ws.cell(row=row, column=7, value=f).border = thin_border

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 10
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 10
ws.column_dimensions['G'].width = 75

wb.save(output_path)
print("Done")
