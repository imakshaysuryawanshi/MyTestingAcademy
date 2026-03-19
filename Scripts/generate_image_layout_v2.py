import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

output_path = r"w:\The Testing Acedamy\Antigravity\My Projects\Project_01_VWO\Test_ Delivarables_Outcome\VWO_Test_Metrics.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Test Metrics"

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def apply_border(cell):
    cell.border = thin_border

# TABLE 1
ws.merge_cells('A1:B1')
ws['A1'] = "Test Metrics"
ws['A1'].font = Font(bold=True, color="FFFFFF", size=12)
ws['A1'].fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid") # Light Blue
ws['A1'].alignment = Alignment(horizontal="center")
apply_border(ws['A1'])
apply_border(ws['B1'])

# Headers for Table 1
ws['A2'] = "Category"
ws['A2'].font = Font(bold=True)
ws['A2'].fill = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid") # Orange
apply_border(ws['A2'])

ws['B2'] = "Count"
ws['B2'].font = Font(bold=True)
ws['B2'].fill = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid") # Orange
apply_border(ws['B2'])

table1_data = [
    ("Number of Requirements", 10),
    ("Average Number of Test Cases Written Per Requirement", 2.7),
    ("Total Number of Test Cases Written for All Requirements", 27),
    ("Total Number of Test Cases Executed", 27),
    ("Number of Test Cases Passed", 12),
    ("Number of Test Cases Failed", 15),
    ("Number of Test Cases Blocked", 0),
    ("Number of Test Cases Unexecuted", 2),
    ("Accessibility", 1),
    ("Functional", 9),
    ("Integration", 1),
    ("Negative", 1),
    ("Performance", 1),
    ("Security", 2)
]

for i, (cat, val) in enumerate(table1_data):
    row = i + 3
    ws.cell(row=row, column=1, value=cat).border = thin_border
    ws.cell(row=row, column=2, value=val).border = thin_border

# Colors for Pass/Fail. Passed is row 7, Failed is row 8 (3+4=7, 3+5=8)
ws['B7'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Green
ws['B8'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Red


# TABLE 2
ws.merge_cells('D4:E4')
ws['D4'] = "Test Metrics"
ws['D4'].font = Font(bold=True, color="FFFFFF", size=12)
ws['D4'].fill = PatternFill(start_color="9900FF", end_color="9900FF", fill_type="solid") # Purple
ws['D4'].alignment = Alignment(horizontal="center")
apply_border(ws['D4'])
apply_border(ws['E4'])

# Headers for Table 2
ws['D5'] = "Category"
ws['D5'].font = Font(bold=True)
ws['D5'].fill = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid") # Orange
apply_border(ws['D5'])

# "Count" or adjusted name - requested "same colmn to second table" but "adjust colmn name correctly" -> "Percentage"
ws['E5'] = "Percentage (%)"
ws['E5'].font = Font(bold=True)
ws['E5'].fill = PatternFill(start_color="F6B26B", end_color="F6B26B", fill_type="solid") # Orange
apply_border(ws['E5'])

table2_data = [
    ("Percentage of Test Cases Executed", 100),
    ("Percentage of Test Cases Not Executed", 0),
    ("Percentage of Test Cases Passed", 44.4),
    ("Percentage of Test Cases Failed", 55.6),
    ("Percentage of Test Cases Blocked", 0)
]

for i, (cat, val) in enumerate(table2_data):
    row = 6 + i
    ws.cell(row=row, column=4, value=cat).border = thin_border
    ws.cell(row=row, column=5, value=val).border = thin_border

# FORMULAS
formulas = [
    "% of Test cases Executed:",
    "(Number of Test Cases Executed / Total Number of Test Cases Written ) * 100",
    "▪ % of test cases NOT executed:",
    "(Number of Test Cases Unexecuted / Total Number of Test Cases Written) * 100",
    "▪ % Test cases passed",
    "(Number of Test Cases Passed / Total Test Cases Executed) * 100",
    "▪ % Test cases failed",
    "(Number of Test Cases Failed / Total Test Cases Executed) * 100",
    "▪ %Test cases blocked",
    "(Number of test cases blocked / Total Test cases executed ) * 100"
]

for i, f in enumerate(formulas):
    row = 6 + i
    ws.cell(row=row, column=7, value=f).border = thin_border

ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 18
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 18
ws.column_dimensions['G'].width = 75

wb.save(output_path)
print("Done v2")
