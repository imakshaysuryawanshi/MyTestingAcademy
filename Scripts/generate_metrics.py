import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Metrics dictionary with corrected spelling
metrics_data = [
    ("Number of Requirements", 10.0),
    ("Average Number of Test Cases Written Per Requirement", 2.0),
    ("Total Number of Test Cases Written for All Requirements", 20.0),
    ("Total Number of Test Cases Executed", 18.0),
    ("Percentage of Test Cases Executed", "90.00%"),
    ("Number of Test Cases Not Executed", 2.0),
    ("Percentage of Test Cases Not Executed", "10.00%"),
    ("Number of Test Cases Passed", 16.0),
    ("Percentage of Test Cases Passed", "88.89%"),
    ("Number of Test Cases Failed", 2.0),
    ("Percentage of Test Cases Failed", "12.50%"),
    ("Number of Test Cases Blocked", 0.0),
    ("Percentage of Test Cases Blocked", "0.00%"),
    ("Total Number of Defects Identified", 10.0),
    ("Critical Defects Count", 2.0),
    ("High Defects Count", 4.0),
    ("Medium Defects Count", 2.0),
    ("Low Defects Count", 2.0),
    ("Customer Defects", 0.0),
    ("Number of Defects Found in UAT", 0.0)
]

output_path = r"w:\The Testing Acedamy\Antigravity\My Projects\Project_01_VWO\Test_ Delivarables_Outcome\VWO_Test_Metrics.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Test Metrics"

# First line: Title
ws['A1'] = "Test Metrics"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:B1')
ws['A1'].alignment = Alignment(horizontal='center')

# Second line: Headers (Category, Count)
ws['A2'] = "Category"
ws['A2'].font = Font(bold=True)
ws['B2'] = "Count"
ws['B2'].font = Font(bold=True)

# Add data
for row_idx, (category, count) in enumerate(metrics_data, start=3):
    ws.cell(row=row_idx, column=1, value=category)
    ws.cell(row=row_idx, column=2, value=count)

# Auto-adjust column widths
ws.column_dimensions['A'].width = 55
ws.column_dimensions['B'].width = 15

wb.save(output_path)
print(f"Metrics file successfully generated: {output_path}")
