import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font

input_path = r"w:\The Testing Acedamy\Antigravity\My Projects\Project_01_VWO\Test_ Delivarables_Outcome\VWO_Test_Case.xlsx"
output_path = r"w:\The Testing Acedamy\Antigravity\My Projects\Project_01_VWO\Test_ Delivarables_Outcome\VWO_Test_Case_Execution.xlsx"

# Read the file
df = pd.read_excel(input_path)

# Add random Status and empty Comment columns
statuses = ["Pass", "Fail"]
df["Status"] = [random.choice(statuses) for _ in range(len(df))]
df["Comment"] = ""

# Save using openpyxl engine to apply styles
df.to_excel(output_path, index=False, engine='openpyxl')

# Load the workbook to apply bold formatting to headers
wb = load_workbook(output_path)
ws = wb.active

# Apply Bold to the first row (headers)
bold_font = Font(bold=True)
for cell in ws[1]:
    cell.font = bold_font

# Save the workbook
wb.save(output_path)
print(f"Execution file successfully generated: {output_path}")
